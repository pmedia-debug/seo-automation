#!/usr/bin/env python3
"""
SEO on-page tag automation script (Excel input).

Usage:
  python seo_automation.py input.xlsx
  python seo_automation.py input.xlsx --xlsx output.xlsx
"""

import json
import sys
from datetime import datetime, timezone, date
from typing import Any, Dict, List, Optional
from urllib.parse import urlparse

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    Workbook = None
    load_workbook = None


def _json_default(obj):
    if isinstance(obj, (datetime, date)):
        return obj.isoformat()
    return str(obj)


def choose_h1(h1: Optional[str], keywords: Optional[str]) -> str:
    if h1 and str(h1).strip():
        return str(h1).strip()
    if keywords and str(keywords).strip():
        return str(keywords).split(",")[0].strip()
    raise ValueError("Missing H1 and keywords; provide at least one.")


def truncate_no_midword(text: str, max_len: int) -> str:
    if len(text) <= max_len:
        return text
    if max_len <= 1:
        return "…"[:max_len]
    limit = max_len - 1
    trimmed = text[:limit].rstrip()
    if " " in trimmed:
        trimmed = trimmed.rsplit(" ", 1)[0]
    if not trimmed:
        trimmed = text[:limit].rstrip()
    return f"{trimmed}…"


def build_meta_title(h1: str, site_name: Optional[str] = None) -> str:
    base = h1.strip()
    if site_name:
        candidate = f"{base} | {str(site_name).strip()}"
        if len(candidate) <= 60:
            return candidate
    return truncate_no_midword(base, 60)


def build_meta_description(h1: str) -> str:
    desc = f"Learn all about {h1} and how it can benefit you."
    return truncate_no_midword(desc, 160)


def titleize_segment(segment: str) -> str:
    words = segment.replace("-", " ").split()
    return " ".join(w.capitalize() for w in words)


def build_breadcrumb_schema(url: str) -> str:
    parsed = urlparse(url)
    if not parsed.scheme or not parsed.netloc:
        raise ValueError(f"Invalid URL: {url}")
    base = f"{parsed.scheme}://{parsed.netloc}/"
    segments = [s for s in parsed.path.split("/") if s]

    items = [
        {
            "@type": "ListItem",
            "position": 1,
            "name": "Home",
            "item": base,
        }
    ]

    current = f"{parsed.scheme}://{parsed.netloc}"
    for i, seg in enumerate(segments, start=2):
        current = f"{current}/{seg}"
        items.append(
            {
                "@type": "ListItem",
                "position": i,
                "name": titleize_segment(seg),
                "item": current,
            }
        )

    data = {
        "@context": "https://schema.org",
        "@type": "BreadcrumbList",
        "itemListElement": items,
    }
    return wrap_json_ld(data)


def build_faq_schema(row: Dict[str, Any]) -> str:
    main_entity = []
    for i in range(1, 6):
        q = row.get(f"faq_q{i}")
        a = row.get(f"faq_a{i}")
        if q and a:
            main_entity.append(
                {
                    "@type": "Question",
                    "name": str(q),
                    "acceptedAnswer": {
                        "@type": "Answer",
                        "text": str(a),
                    },
                }
            )
    if not main_entity:
        return ""
    data = {
        "@context": "https://schema.org",
        "@type": "FAQPage",
        "mainEntity": main_entity,
    }
    return wrap_json_ld(data)


def build_product_schema(row: Dict[str, Any]) -> str:
    required = [
        "product_name",
        "product_url",
        "product_image",
        "brand_name",
        "product_description",
        "rating_value",
        "best_rating",
    ]
    for key in required:
        if not row.get(key):
            raise ValueError(f"Missing required product field: {key}")

    data = {
        "@context": "https://schema.org",
        "@type": "Product",
        "name": row["product_name"],
        "url": row["product_url"],
        "image": row["product_image"],
        "brand": {
            "@type": "Organization",
            "name": row["brand_name"],
        },
        "description": row["product_description"],
        "review": {
            "@type": "Review",
            "reviewRating": {
                "@type": "Rating",
                "ratingValue": str(row["rating_value"]),
                "bestRating": str(row["best_rating"]),
            },
            "author": {
                "@type": "Organization",
                "name": row.get("review_author_name", row["brand_name"]),
            },
        },
    }
    return wrap_json_ld(data)


def build_blog_schema(row: Dict[str, Any]) -> str:
    required = [
        "blog_url",
        "headline",
        "blog_description",
        "blog_image",
        "publisher_name",
        "publisher_logo",
        "author_name",
    ]
    for key in required:
        if not row.get(key):
            raise ValueError(f"Missing required blog field: {key}")

    date_published = row.get("date_published")
    date_modified = row.get("date_modified")
    if not date_published or not date_modified:
        today = datetime.now(timezone.utc).date().isoformat()
        date_published = date_published or today
        date_modified = date_modified or today

    data = {
        "@context": "https://schema.org",
        "@type": "Article",
        "mainEntityOfPage": {
            "@type": "WebPage",
            "@id": row["blog_url"],
        },
        "headline": row["headline"],
        "description": row["blog_description"],
        "image": row["blog_image"],
        "author": {
            "@type": "Organization",
            "name": row["author_name"],
        },
        "publisher": {
            "@type": "Organization",
            "name": row["publisher_name"],
            "logo": {
                "@type": "ImageObject",
                "url": row["publisher_logo"],
            },
        },
        "datePublished": date_published,
        "dateModified": date_modified,
    }
    return wrap_json_ld(data)


def wrap_json_ld(data: Dict[str, Any]) -> str:
    json_text = json.dumps(data, indent=2, ensure_ascii=False, default=_json_default)
    return f'<script type="application/ld+json">{json_text}</script>'


def build_outputs(row: Dict[str, Any]) -> Dict[str, str]:
    h1 = choose_h1(row.get("h1"), row.get("keywords"))
    meta_title = build_meta_title(h1, row.get("site_name"))
    meta_description = build_meta_description(h1)

    if not row.get("url"):
        raise ValueError("Missing required field: url")

    breadcrumb_schema = build_breadcrumb_schema(row["url"])
    product_schema = build_product_schema(row)
    faq_schema = build_faq_schema(row)
    blog_schema = build_blog_schema(row)

    return {
        "url": row["url"],
        "meta_title": meta_title,
        "meta_description": meta_description,
        "product_schema": product_schema,
        "breadcrumb_schema": breadcrumb_schema,
        "faq_schema": faq_schema,
        "blog_schema": blog_schema,
    }


def read_excel(path: str) -> List[Dict[str, Any]]:
    if load_workbook is None:
        raise ValueError("openpyxl is not installed. Run: python3 -m pip install openpyxl")

    wb = load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel file is empty.")

    headers = [str(h).strip() if h else "" for h in rows[0]]
    data_rows = []
    for r in rows[1:]:
        row_dict = {}
        for idx, val in enumerate(r):
            key = headers[idx] if idx < len(headers) else ""
            if key:
                row_dict[key] = val
        if any(v is not None and str(v).strip() != "" for v in row_dict.values()):
            data_rows.append(row_dict)
    return data_rows


def write_xlsx(rows: List[Dict[str, str]], out_path: str) -> None:
    if Workbook is None:
        raise ValueError("openpyxl is not installed. Run: python3 -m pip install openpyxl")

    headers = [
        "url",
        "meta_title",
        "meta_description",
        "product_schema",
        "breadcrumb_schema",
        "faq_schema",
        "blog_schema",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "SEO Output"
    ws.append(headers)

    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    wb.save(out_path)


def main() -> int:
    if len(sys.argv) not in (2, 4):
        print("Usage: python seo_automation.py input.xlsx [--xlsx output.xlsx]")
        return 1

    xlsx_path = None
    if len(sys.argv) == 4:
        if sys.argv[2] != "--xlsx":
            print("Usage: python seo_automation.py input.xlsx [--xlsx output.xlsx]")
            return 1
        xlsx_path = sys.argv[3]

    try:
        rows = read_excel(sys.argv[1])

        outputs_all = []
        for i, row in enumerate(rows, start=1):
            outputs = build_outputs(row)
            outputs_all.append(outputs)

            print(f"=== Page {i} ===")
            print(f"Meta Title: {outputs['meta_title']}")
            print(f"Meta Description: {outputs['meta_description']}\n")
            print("Product Schema:")
            print(outputs["product_schema"])
            print()
            print("Breadcrumb Schema:")
            print(outputs["breadcrumb_schema"])
            print()
            if outputs["faq_schema"]:
                print("FAQ Schema:")
                print(outputs["faq_schema"])
                print()
            print("Blog Post Schema:")
            print(outputs["blog_schema"])
            print()

        if xlsx_path:
            write_xlsx(outputs_all, xlsx_path)
            print(f"XLSX saved to: {xlsx_path}")

        return 0
    except ValueError as exc:
        print(f"Error: {exc}")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
